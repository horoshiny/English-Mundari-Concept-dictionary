
import pandas as pd
from openpyxl import load_workbook
import xlwt
import xlsxwriter
import re

path1="C:\\Users\\horos\\Documents\\mundari-eng dictionary\\concept_dictionary_user.xlsx"


df1=pd.read_excel('C:\\Users\\horos\\Documents\\mundari-eng dictionary\\concept_dictionary_user.xlsx')

wb=load_workbook(path1)

ws=wb['Sheet1']
ws.cell(row=1, column=9).value=("Mundari")

all_rows=list(ws.rows)
wb.save(path1)
print(f"Found {len(all_rows)} rows of data.")

f1 = open("C:\\Users\\horos\\Documents\\mundari-eng dictionary\\1.txt", "r")
f2 = open("C:\\Users\\horos\\Documents\\mundari-eng dictionary\\munda.txt", "r",encoding="utf-8")


f1_data = f1.readlines()
f2_data = f2.readlines()
dictionary={}
i = 0
for line1 in f1_data:
    for word1 in line1.split(","):
        k=1
        for line2 in f2_data:
            i += 1
            for word2 in line2.split():
                if word1 == word2:
                    print("**********************")
                    print("Line ", i, ": IDENTICAL")
                    print("Word:", word1)
                    print("Line ", i, ":", line2)
                    final_word1=line2.split(';')[0]
                    final_word=final_word1.split(';')
                    print(final_word)
                    final_word=final_word1.split('—')[0]
                    if ((df1.loc[df1['Eng'] == word1, 'Mundari'])).empty:
                        final_word = final_word + str('_') + str(k)  # adding numbers at the end
                        word1 = word1 + str('_') + str(k)
                        print("WORD1: ", word1)
                        print("Final word: ", final_word)
                        df1.loc[df1['Eng'] == word1, 'Mundari'] = final_word                 
                    else:
                        k += 1
                        final_word = final_word + str('_') + str(k)  # adding numbers at the end
                        word1 = word1 + str('_') + str(k)
                        df1.loc[df1['Eng']==word1,'Mundari']=final_word

                        print("value of k: ",k)


                    df1.to_excel('C:\\Users\\horos\\Documents\\mundari-eng dictionary\\concept_dictionary_user.xlsx', index=None)


print(df1)
print("All done!")

# closing files
f1.close()
f2.close()
